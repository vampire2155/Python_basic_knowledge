import json
from APITest.operateExcel.login0616 import login0616
from openpyxl import load_workbook   #openpyxl操作excel下标都是以 1 开始，而不是 以 0开始
#使用 openpyxl 当中的 load_workbook 来处理已存在的 Excel 文件数据，只能读写以 .xlsx 为拓展名的文件。

excelDir = r'G:\测试学习资料\接口测试\接口自动化测试_6-15\用例.xlsx'
# 第一步. 打开excel文件：使用load_workbook()传入  文件名（py文件和excel文件在同一个文件夹内）   或者  文件路径和文件名（py文件和excel文件 不在 同一个文件夹内）
wb = load_workbook(excelDir)  # 返回创建一个Workbook的对象, 相当是一个excel文件

# 第二步. 定位表单两种方式
# ws = wb.active   #方式一：默认获取第一个激活的表单, 会创建一个Worksheet对象, 相当于一个表单
ws = wb['Sheet1']  #方式二：指定表单
# new_workBook = copy(wb)
# new_workSheet = new_workBook.get_sheet(1)  #通过下标获取sheet页
'''
# 第三步. 定位单元格 cell
cellData = ws.cell(row=2,column=2)     # 会创建一个Cell对象, 相当于一个单元格

# 使用 Cell()对象中的value属性, 能获取单元格中的值
print(cellData.value)

#单元格中写入数据的两种方式
# 方法一: 定位单元格后，使用value属性，将数据写入到指定的单元格
cellData.value = "休想"   # 修改单元格的值

# 方法二: 定位表单，使用cell方法，将数据写入到指定的单元格
ws.cell(row=2, column=3, value="休想")

#保存excel
wb.save("cases.xlsx")
'''
def listClass_OperExcel(row):
    cellData = json.loads(ws.cell(row=row,column=7).value)  #获取第2行，第7列中的值，并转换成 字典类型。
    usrName = cellData['username']
    psd = cellData['password']
    EXEresult = login0616(username=usrName,password=psd)['retcode']   #调用了login0616()函数然并返回用例的执行结果
    tcNum = ws.cell(row=row,column=1).value     #获取excel中的用例编号
    expected_result = json.loads(ws.cell(row=row,column=9).value)['retcode']  #获取用例中的返回结果列中的值
    if EXEresult == expected_result:
        print(f'--------------用例编号为：{tcNum}的用例执行pass----------------')
        info = 'pass'  # 用例执行的结果赋值给一个变量，会在下面用例结果中写入时用到。
    else:
        print(f'--------------用例编号为：{tcNum}的用例执行fail-----------------')
        info = 'fail'  # 用例执行的结果赋值给一个变量，会在下面用例结果中写入时用到。
    ws.cell(row=row,column=10,value=info)  #(row=行,column=列,value=内容)    #把执行结果写入到单元格中。

for row in range(2,6):
    listClass_OperExcel(row)
wb.save(r'G:\测试学习资料\接口测试\接口自动化测试_6-15\xlsx_result.xlsx')  #保存执行结果。
#经验证，使用openpyxl 模块的 load_workbook 方法操作以后缀名为 xlsx 的excel时，如果需要把用例的执行结果保存在另外一个excel中时，
#直接在保存时 另外命名一个 excel即可。不需要 进行复制。





