#获取excel中的用例并执行，最后把执行结果 回填到excel对应的列中。
import json,requests
from openpyxl import load_workbook
#把登录接口封装成一个函数
def loginAPI(username='auto',password='sdfsdfsdf'):
    HOST = 'http://127.0.0.1:9000'
    url = f'{HOST}/api/mgr/loginReq'
    form_data = {
                'username':username,
                'password':password
                }
    res = requests.post(url,data=form_data)
    res.encoding = 'unicode-escape'
    return res.json()

excelDir = r'G:\测试学习资料\接口测试\接口自动化测试_6-15\用例.xlsx'
#打开excel文件
workBook = load_workbook(excelDir)   #会创建一个workbook对象，是一个excel文件
#定位表单 （即是哪一个sheet页）
workSheet = workBook['Sheet1']
#iter_rows(min_row=2, max_row=2, values_only=True)  min_row=2, max_row=2 表示取第二行的所有数据 .
# values_only=True表示获取的是单元格的值  values_only=False 表示获取的是单元格的对象 ，例如 <Cell '用例'.A2>
columnData = tuple(workSheet.iter_rows(min_col=7,max_col=7,values_only=True))[1:5]
list = []
for one in columnData:
    usrNmae, psd = json.loads(one[0])["username"],json.loads(one[0])["password"]
    EXEresult = loginAPI(username=usrNmae, password=psd)['retcode']
    list.append(EXEresult)
for row in range(2,6):
    tcNum = workSheet.cell(row=row,column=1).value   #cell(row=row,column=1) 行和列的中必须至少为 1
    expected_result = json.loads(workSheet.cell(row=row,column=9).value)['retcode']
    print (expected_result)
    if list[row-2] == expected_result:  #因为list的范围是 0--4 ，而row的范围是 2--6 ，所以需要减去 2
        print(f'--------------用例编号为：{tcNum}的用例执行pass----------------')
        info = 'pass'  # 用例执行的结果赋值给一个变量，会在下面用例结果中写入时用到。
    else:
        print(f'--------------用例编号为：{tcNum}的用例执行fail-----------------')
        info = 'fail'  # 用例执行的结果赋值给一个变量，会在下面用例结果中写入时用到。
    workSheet.cell(row=row, column=10, value=info)  # (row=行,column=列,value=内容)    #把执行结果写入到单元格中。

workBook.save(r'G:\测试学习资料\接口测试\接口自动化测试_6-15\iter_rows_result.xlsx')





