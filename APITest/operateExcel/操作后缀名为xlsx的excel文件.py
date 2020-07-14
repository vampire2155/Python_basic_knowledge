#获取excel中的用例并执行，最后把执行结果 回填到excel对应的列中。
import json,requests
from openpyxl import load_workbook
#把登录接口封装成一个函数
def loginAPI(username='auto',password='sdfsdfsdf'):
    HOST = 'http://127.0.0.1:9000'
    url = f'{HOST}/api/mgr/loginReq'
    form_data = {
                'username':username,
                'password':password
                }
    res = requests.post(url,data=form_data)
    res.encoding = 'unicode-escape'
    return res.json()

#使用openpyxl 的 load_workbook 方法操作后缀名为 xlsx的excel文件并写入结果。
excelDir = r'G:\测试学习资料\接口测试\接口自动化测试_6-15\用例.xlsx'
wb = load_workbook(excelDir)  # 返回创建一个Workbook的对象, 相当是一个excel文件
ws = wb['Sheet1']  #方式二：指定表单
def listClass_OperExcel(row):
    cellData = json.loads(ws.cell(row=row,column=7).value)  #获取第2行，第7列中的值，并转换成 字典类型。
    usrName = cellData['username']
    psd = cellData['password']
    EXEresult = loginAPI(username=usrName,password=psd)['retcode']   #调用了login0616()函数然并返回用例的执行结果
    tcNum = ws.cell(row=row,column=1).value     #获取excel中的用例编号
    expected_result = json.loads(ws.cell(row=row,column=9).value)['retcode']  #获取用例中的返回结果列中的值
    if EXEresult == expected_result:
        print(f'--------------用例编号为：{tcNum}的用例执行pass----------------')
        info = 'pass'  # 用例执行的结果赋值给一个变量，会在下面用例结果中写入时用到。
    else:
        print(f'--------------用例编号为：{tcNum}的用例执行fail-----------------')
        info = 'fail'  # 用例执行的结果赋值给一个变量，会在下面用例结果中写入时用到。
    ws.cell(row=row,column=10,value=info)  #(row=行,column=列,value=内容)    #把执行结果写入到单元格中。

for row in range(2,6):
    listClass_OperExcel(row)
wb.save(r'G:\测试学习资料\接口测试\接口自动化测试_6-15\xlsx_result.xlsx')  #保存执行结果。
#经验证，使用openpyxl 模块的 load_workbook 方法操作以后缀名为 xlsx 的excel时，如果需要把用例的执行结果保存在另外一个excel中时，
#直接在保存时 另外命名一个 excel即可。不需要 进行复制。








